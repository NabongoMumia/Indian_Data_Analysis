import os
import io
import json
import psycopg2
from psycopg2.extras import RealDictCursor
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pypdf import PdfWriter
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing, String, PolyLine
from reportlab.graphics.charts.barcharts import VerticalBarChart

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "fallback-secret-key-change-me")

SUBJECTS = ['English', 'Maths', 'Chemistry', 'Biology', 'ICT', 'Physics', 'Business', 'Geography', 'Islamiyat']


def get_db_connection():
    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL environment variable is missing.")
    
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
        
    return psycopg2.connect(db_url, cursor_factory=RealDictCursor, sslmode='require')


def init_db():
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL
            );
        ''')
        
        cur.execute('''
            CREATE TABLE IF NOT EXISTS students (
                adm_no VARCHAR(50) PRIMARY KEY,
                student_name VARCHAR(150) NOT NULL,
                class_name VARCHAR(50) NOT NULL,
                assessment_term VARCHAR(100),
                opening_date VARCHAR(20),
                closing_date VARCHAR(20),
                class_teacher VARCHAR(100),
                principal_name VARCHAR(100),
                english FLOAT DEFAULT 0,
                maths FLOAT DEFAULT 0,
                chemistry FLOAT DEFAULT 0,
                biology FLOAT DEFAULT 0,
                ict FLOAT DEFAULT 0,
                physics FLOAT DEFAULT 0,
                business FLOAT DEFAULT 0,
                geography FLOAT DEFAULT 0,
                islamiyat FLOAT DEFAULT 0,
                teachers JSONB DEFAULT '{}'::jsonb
            );
        ''')
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"Database initialization error: {e}")
    finally:
        if conn:
            conn.close()


with app.app_context():
    init_db()


def get_grade_and_remark(score):
    if score >= 80:
        return 'A*', 'Excellent'
    elif score >= 70:
        return 'A', 'Very Good'
    elif score >= 60:
        return 'B', 'Good'
    elif score >= 50:
        return 'C', 'Satisfactory'
    elif score >= 40:
        return 'D', 'Pass'
    else:
        return 'E', 'Needs Improvement'


def create_student_pdf_buffer(student):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=20,
        bottomMargin=20
    )
    styles = getSampleStyleSheet()
    
    NAVY = colors.HexColor('#0A2540')
    HEADER_BLUE = colors.HexColor('#1E3A8A')
    LIGHT_BG = colors.HexColor('#F8FAFC')
    LINE_ORANGE = colors.HexColor('#D97706')

    title_style = ParagraphStyle(
        'HeaderTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=NAVY,
        alignment=0
    )
    
    sub_title_style = ParagraphStyle(
        'HeaderSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#475569'),
        alignment=0
    )
    
    motto_style = ParagraphStyle(
        'HeaderMotto',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#1E40AF'),
        alignment=0
    )
    
    section_heading = ParagraphStyle(
        'SecHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
        textColor=NAVY,
        alignment=1
    )

    elements = []

    # 1. Top Header with Logo inserted on Left
    logo_path = os.path.join(app.root_path, 'lmage1.png')
    if not os.path.exists(logo_path):
        logo_path = os.path.join(app.root_path, 'static', 'lmage1.png')

    header_text_cells = [
        Paragraph("SKY INTERNATIONAL SCHOOLS", title_style),
        Paragraph("Email: info@skyschools.net | Hargeisa, Somaliland", sub_title_style),
        Paragraph("'No Substitute For Self Discipline'", motto_style)
    ]

    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=65, height=65)
        header_table = Table([[logo_img, header_text_cells]], colWidths=[75, 465])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('PADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(header_table)
    else:
        title_style.alignment = 1
        sub_title_style.alignment = 1
        motto_style.alignment = 1
        for item in header_text_cells:
            elements.append(item)

    elements.append(Spacer(1, 6))
    elements.append(Paragraph("SECONDARY PROGRESS & ACHIEVEMENT REPORT", section_heading))
    elements.append(Spacer(1, 8))

    # 2. Metadata Block
    meta_data = [
        [f"Student Name: {student['student_name']}", f"Admission No: {student['adm_no']}"],
        [f"Class: {student['class_name']}", f"Assessment: {student.get('assessment_term', 'N/A')}"],
        [f"Opening Date: {student.get('opening_date', 'N/A')}", f"Closing Date: {student.get('closing_date', 'N/A')}"]
    ]
    t_meta = Table(meta_data, colWidths=[270, 270])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), LIGHT_BG),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
        ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold')
    ]))
    elements.append(t_meta)
    elements.append(Spacer(1, 10))
    
    # --- ADD TITLE HERE ---
    table_title_style = ParagraphStyle(
        'TableTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=13,
        textColor=NAVY,
        alignment=0  # Left-aligned (use 1 for centered if preferred)
    )
    elements.append(Paragraph("Official Academic Progress", table_title_style))
    elements.append(Spacer(1, 4))
    # ----------------------
    
    # 3. Subject Performance Table
    score_data = [["Subject", "Total Marks", "Marks Obtained", "Subject Grade", "Remarks", "Teacher"]]
    teachers = student.get("teachers") or {}
    if isinstance(teachers, str):
        try:
            teachers = json.loads(teachers)
        except Exception:
            teachers = {}

    total_score = 0
    count = 0
    scores_list = []
    grades_list = []

    for sub in SUBJECTS:
        score = student.get(sub.lower(), 0.0)
        total_score += score
        count += 1
        scores_list.append(score)
        
        grade, remark = get_grade_and_remark(score)
        grades_list.append(grade)
        t_name = teachers.get(sub, "")
        score_data.append([sub, "100", f"{score:.1f}", grade, remark, t_name])

    avg_score = total_score / count if count > 0 else 0
    overall_grade, _ = get_grade_and_remark(avg_score)
    
    score_data.append(["TOTAL MARKS", f"{total_score:.1f} / {count * 100}", "", "", "", ""])
    score_data.append(["AVERAGE PERCENTAGE", f"{avg_score:.2f}%", f"FINAL GRADE: {overall_grade}", "", "", ""])

    t_scores = Table(score_data, colWidths=[90, 75, 95, 80, 100, 100])
    t_scores.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HEADER_BLUE),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#94A3B8')),
        ('BACKGROUND', (0,-2), (-1,-1), colors.HexColor('#E2E8F0')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,-2), (-1,-1), 'Helvetica-Bold'),
        ('SPAN', (1,-2), (5,-2)),
        ('SPAN', (2,-1), (5,-1)),
        ('PADDING', (0,0), (-1,-1), 3.5),
        ('FONTSIZE', (0,0), (-1,-1), 8)
    ]))
    elements.append(t_scores)
    elements.append(Spacer(1, 8))

    # 4. Grading System Key
    scale_data = [
        ["Grading System", "80-100", "70-79", "60-69", "50-59", "40-49", "0-39"],
        ["Grade Scale", "A*", "A", "B", "C", "D", "E"]
    ]
    t_scale = Table(scale_data, colWidths=[90] + [75]*6)
    t_scale.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0,0), (-1,0), LIGHT_BG),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 7.5),
        ('PADDING', (0,0), (-1,-1), 3)
    ]))
    elements.append(t_scale)
    elements.append(Spacer(1, 8))

    # 5. Performance Chart
    elements.append(Paragraph("Subject Performance Analysis & Progress View", ParagraphStyle('ChartHeading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, alignment=1, textColor=NAVY)))
    elements.append(Spacer(1, 4))

    drawing_w, drawing_h = 540, 115
    drawing = Drawing(drawing_w, drawing_h)
    
    bc = VerticalBarChart()
    bc.x = 35
    bc.y = 18
    bc.height = 75
    bc.width = 480
    bc.data = [scores_list]
    bc.categoryAxis.categoryNames = SUBJECTS
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.dy = -10
    bc.categoryAxis.labels.fontName = 'Helvetica'
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = 100
    bc.valueAxis.valueStep = 20
    bc.valueAxis.labels.fontSize = 7
    bc.bars[0].fillColor = colors.HexColor('#2563EB')
    bc.bars[0].strokeColor = colors.HexColor('#1D4ED8')
    bc.bars[0].strokeWidth = 0.5
    
    drawing.add(bc)

    n_bars = len(scores_list)
    bar_group_width = bc.width / float(n_bars)
    points = []
    
    for i, (score, grade) in enumerate(zip(scores_list, grades_list)):
        center_x = bc.x + (i + 0.5) * bar_group_width
        val_clamped = min(max(score, 0), 100)
        center_y = bc.y + (val_clamped / 100.0) * bc.height
        points.extend([center_x, center_y])

        label_y = min(center_y + 4, bc.y + bc.height + 2)
        drawing.add(String(center_x, label_y, grade, fontName='Helvetica-Bold', fontSize=7.5, textAnchor='middle', fillColor=colors.HexColor('#991B1B')))

    if len(points) >= 4:
        drawing.add(PolyLine(points, strokeColor=LINE_ORANGE, strokeWidth=1.5))

    elements.append(drawing)
    elements.append(Spacer(1, 10))

    # 6. Teacher Remarks & Signatures Section
    remark_heading_style = ParagraphStyle(
        'RemarkHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=13,
        textColor=NAVY,
        alignment=1
    )
    
    remark_text_style = ParagraphStyle(
        'RemarkText',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1E293B')
    )

    remark_text = "Very good effort. Consistently strong results this term." if avg_score >= 70 else "Good results, but with additional focus higher scores are reachable."
    
    teacher_sig = student.get('class_teacher') or "None"
    principal_sig = student.get('principal_name') or "None"

    elements.append(Paragraph("Teacher Remarks", remark_heading_style))
    elements.append(Spacer(1, 4))

    # Light background boxed container for teacher remark
    box_data = [[Paragraph(f'"{remark_text}"', remark_text_style)]]
    t_box = Table(box_data, colWidths=[540])
    t_box.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F1F5F9')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t_box)
    elements.append(Spacer(1, 16))

    # Clean 2-column signature layout matching the target image
    sig_line = "_____________________________________"
    sig_data = [
        [
            Paragraph(f"<b>Signature:</b> {sig_line}", styles['Normal']),
            Paragraph(sig_line, styles['Normal'])
        ],
        [
            Paragraph(f"<b>Class Teacher:</b> {teacher_sig}", styles['Normal']),
            Paragraph(f"<b>Principal:</b> {principal_sig}", styles['Normal'])
        ]
    ]

    t_sig = Table(sig_data, colWidths=[270, 270])
    t_sig.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 2),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    elements.append(t_sig)

    doc.build(elements)
    buffer.seek(0)
    return buffer


@app.route("/")
def index():
    if "username" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/login", methods=["POST"])
def login():
    action = request.form.get("action")
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()

    if not username or not password:
        flash("Username and password are required.", "danger")
        return redirect(url_for("index"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        if action == "register":
            cur.execute("SELECT id FROM users WHERE username = %s;", (username,))
            if cur.fetchone():
                flash("Username already exists.", "warning")
            else:
                hashed_pw = generate_password_hash(password)
                cur.execute("INSERT INTO users (username, password) VALUES (%s, %s);", (username, hashed_pw))
                conn.commit()
                session["username"] = username
                flash("Account registered successfully!", "success")
                return redirect(url_for("dashboard"))

        elif action == "login":
            cur.execute("SELECT * FROM users WHERE username = %s;", (username,))
            user = cur.fetchone()
            
            if user and check_password_hash(user["password"], password):
                session["username"] = username
                flash("Logged in successfully!", "success")
                return redirect(url_for("dashboard"))
            else:
                flash("Invalid credentials.", "danger")

    except Exception as e:
        conn.rollback()
        print(f"Login/Register Error: {e}")
        flash(f"An unexpected error occurred: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()

    return redirect(url_for("index"))


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("index"))


@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect(url_for("index"))

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT DISTINCT class_name FROM students WHERE class_name IS NOT NULL AND class_name != '';")
        classes = [row["class_name"] for row in cur.fetchall()]
    finally:
        cur.close()
        conn.close()

    return render_template("dashboard.html", username=session["username"], subjects=SUBJECTS, classes=classes)


@app.route("/api/student/<adm_no>")
def get_student(adm_no):
    if "username" not in session:
        return jsonify({"found": False, "error": "Unauthorized"}), 401

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM students WHERE adm_no = %s;", (adm_no,))
        student = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if student:
        student["found"] = True
        return jsonify(dict(student))
    return jsonify({"found": False})


@app.route("/save_student", methods=["POST"])
def save_student():
    if "username" not in session:
        return redirect(url_for("index"))

    adm_no = request.form.get("adm_no")
    student_name = request.form.get("student_name")
    class_name = request.form.get("class_name")
    assessment_term = request.form.get("assessment_term")
    opening_date = request.form.get("opening_date")
    closing_date = request.form.get("closing_date")
    class_teacher = request.form.get("class_teacher")
    principal_name = request.form.get("principal_name")

    teachers = {}
    scores = {}
    for sub in SUBJECTS:
        score_val = request.form.get(f"score_{sub}", 0)
        try:
            scores[sub.lower()] = float(score_val)
        except ValueError:
            scores[sub.lower()] = 0.0

        teachers[sub] = request.form.get(f"teacher_{sub}", "")

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO students (
                adm_no, student_name, class_name, assessment_term, opening_date, closing_date,
                class_teacher, principal_name, english, maths, chemistry, biology, ict,
                physics, business, geography, islamiyat, teachers
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (adm_no) DO UPDATE SET
                student_name = EXCLUDED.student_name,
                class_name = EXCLUDED.class_name,
                assessment_term = EXCLUDED.assessment_term,
                opening_date = EXCLUDED.opening_date,
                closing_date = EXCLUDED.closing_date,
                class_teacher = EXCLUDED.class_teacher,
                principal_name = EXCLUDED.principal_name,
                english = EXCLUDED.english,
                maths = EXCLUDED.maths,
                chemistry = EXCLUDED.chemistry,
                biology = EXCLUDED.biology,
                ict = EXCLUDED.ict,
                physics = EXCLUDED.physics,
                business = EXCLUDED.business,
                geography = EXCLUDED.geography,
                islamiyat = EXCLUDED.islamiyat,
                teachers = EXCLUDED.teachers;
        """, (
            adm_no, student_name, class_name, assessment_term, opening_date, closing_date,
            class_teacher, principal_name, scores.get("english", 0), scores.get("maths", 0),
            scores.get("chemistry", 0), scores.get("biology", 0), scores.get("ict", 0),
            scores.get("physics", 0), scores.get("business", 0), scores.get("geography", 0),
            scores.get("islamiyat", 0), json.dumps(teachers)
        ))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    flash(f"Record for {student_name} saved successfully!", "success")
    return redirect(url_for("dashboard"))


@app.route("/generate_pdf/<adm_no>")
def generate_pdf(adm_no):
    if "username" not in session:
        return redirect(url_for("index"))

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM students WHERE adm_no = %s;", (adm_no,))
        student = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if not student:
        flash("Student record not found.", "danger")
        return redirect(url_for("dashboard"))

    pdf_buffer = create_student_pdf_buffer(student)
    return send_file(pdf_buffer, as_attachment=True, download_name=f"Report_{adm_no}.pdf", mimetype="application/pdf")


@app.route("/download_class_pdf", methods=["POST"])
def download_class_pdf():
    if "username" not in session:
        return redirect(url_for("index"))

    class_name = request.form.get("class_name")
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM students WHERE class_name = %s ORDER BY student_name ASC;", (class_name,))
        students = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    if not students:
        flash("No students found for this class.", "warning")
        return redirect(url_for("dashboard"))

    merger = PdfWriter()
    for st in students:
        st_buffer = create_student_pdf_buffer(st)
        merger.append(st_buffer)

    output_buffer = io.BytesIO()
    merger.write(output_buffer)
    merger.close()
    output_buffer.seek(0)

    return send_file(output_buffer, as_attachment=True, download_name=f"Class_Reports_{class_name}.pdf", mimetype="application/pdf")


@app.route("/export_excel", methods=["POST"])
def export_excel():
    if "username" not in session:
        return redirect(url_for("index"))

    class_name = request.form.get("class_name")
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM students WHERE class_name = %s;", (class_name,))
        students = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    if not students:
        flash("No students found to export.", "warning")
        return redirect(url_for("dashboard"))

    # Calculate totals, averages, and grades first for sorting
    processed_students = []
    for st in students:
        total = 0
        subject_scores = []
        for sub in SUBJECTS:
            val = st.get(sub.lower(), 0.0)
            subject_scores.append(val)
            total += val
        
        avg = total / len(SUBJECTS) if SUBJECTS else 0.0
        grade, _ = get_grade_and_remark(avg)
        
        processed_students.append({
            "adm_no": st["adm_no"],
            "student_name": st["student_name"],
            "class_name": st["class_name"],
            "term": st.get("assessment_term", ""),
            "scores": subject_scores,
            "total": total,
            "avg": round(avg, 2),
            "grade": grade
        })

    # Sort students by Total Score in descending order
    processed_students.sort(key=lambda x: x["total"], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"MARKLIST_{class_name.upper()}"

    headers = ["Adm No", "Student Name", "Class", "Term"] + SUBJECTS + ["Total Score", "Average (%)", "Grade"]
    max_col = len(headers)

    # 1. Main Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = ws.cell(row=1, column=1, value="SKY SCHOOLS INTERNATIONAL")
    title_cell.font = Font(name="Calibri", size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 2. Sub-Header Row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sub_title_cell = ws.cell(row=2, column=1, value=f"MARKLIST_{class_name.upper()}")
    sub_title_cell.font = Font(name="Calibri", size=12, bold=True)
    sub_title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 3. Table Headers Row
    ws.append([]) # Row 3 placement
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_num > 2 else "left", vertical="center")

    # 4. Data Rows
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for row_idx, st in enumerate(processed_students, start=4):
        row_data = [
            st["adm_no"],
            st["student_name"],
            st["class_name"],
            st["term"]
        ] + st["scores"] + [st["total"], st["avg"], st["grade"]]
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            
            if col_idx in [1, 3, 4] or col_idx >= 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"MARKLIST_{class_name}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run(debug=True)